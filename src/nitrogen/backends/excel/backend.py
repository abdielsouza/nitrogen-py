from nitrogen.backends.base import Backend
from nitrogen.backends.excel.compiler import ExcelCompiler
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
from typing import Type
from nitrogen.core import Sheet, Column, Formula

class ExcelBackend(Backend):
    def __init__(self, workbook: Workbook):
        self.__workbook = workbook
        self.__compiler = ExcelCompiler()

    def create_sheet(self, sheet: Type[Sheet]) -> None:
        # create the worksheet for the Sheet subclass
        self.__workbook.create_sheet(title=sheet.default_name())

    def write_column(self, sheet: Type[Sheet], column: Column) -> None:
        ws = self.__workbook[sheet.default_name()]

        # find next empty column in header (row 1)
        col_idx = 1
        while ws.cell(row=1, column=col_idx).value is not None:
            col_idx += 1

        # write header
        ws.cell(row=1, column=col_idx, value=column.name)

        # write existing data rows for this column (if any)
        rows = getattr(sheet, "__rows__", [])
        for row_idx, row in enumerate(rows):
            value = row.get(column.name)
            ws.cell(row=2 + row_idx, column=col_idx, value=value)

    def write_formula(self, sheet: Type[Sheet], formula: Formula) -> None:
        ws = self.__workbook[sheet.default_name()]

        if formula.name is None:
            raise ValueError("formula name cannot be null.")

        compiled = formula.compile(self.__compiler)
        if not isinstance(compiled, str):
            compiled = str(compiled)

        header_map: dict[str, int] = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                header_map[str(cell.value)] = idx

        target_col = header_map.get(formula.name)
        if target_col is None:
            target_col = 1
            while ws.cell(row=1, column=target_col).value is not None:
                target_col += 1

            ws.cell(row=1, column=target_col, value=formula.name)
            header_map[formula.name] = target_col

        names = [name for name in header_map.keys() if name != formula.name]
        names.sort(key=len, reverse=True)

        rows = getattr(sheet, "__rows__", [])
        for row_idx, _ in enumerate(rows):
            row_number = 2 + row_idx
            formula_text = compiled

            for name in names:
                col_idx = header_map.get(name)
                if col_idx is None:
                    continue

                col_letter = get_column_letter(col_idx)
                formula_text = re.sub(
                    r"(?<![A-Za-z0-9_])" + re.escape(name) + r"(?![A-Za-z0-9_])",
                    f"{col_letter}{row_number}",
                    formula_text,
                )
            
            formula_value = formula_text if formula_text.startswith("=") else f"={formula_text}"
            ws.cell(row=row_number, column=target_col, value=formula_value)

    def save(self, path: str | None = None) -> None:
        if path:
            self.__workbook.save(path)
        else:
            raise ValueError("missing path for excel file saving")