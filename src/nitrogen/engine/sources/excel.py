from typing import cast, Any
import openpyxl as xl
import os

from nitrogen.engine.compilers.excel import ExcelCompiler
from nitrogen.engine.contexts import ExcelContext
from nitrogen.engine.query import FetchQuery
from nitrogen.engine.source import DataSource

class ExcelDataSource(DataSource):
    """Implementation of Excel data source."""

    def __init__(self, filepath: str):
        if os.path.exists(filepath):
            self._workbook = xl.load_workbook(filepath)
        else:
            self._workbook = xl.Workbook()

        self._context = ExcelContext()
        self._compiler = ExcelCompiler()
        self._filepath = filepath

    def execute(self, query):
        if query.sheet not in self._workbook.sheetnames:
            self._workbook.create_sheet(query.sheet)

        worksheet = self._workbook[query.sheet]
        self._context.worksheet = worksheet
        result = self._compiler.compile(query, self._context)
        result = cast(Any, result)

        if not isinstance(query, FetchQuery):
            self._workbook.save(self._filepath)

        return result
    
    @property
    def context(self):
        return self._context

    @property
    def workbook(self):
        return self._workbook
    
    @property
    def filepath(self):
        return self._filepath
