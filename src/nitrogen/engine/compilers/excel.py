from collections import defaultdict
from typing import Any, Callable, cast

from nitrogen.engine.compiler import QueryCompiler
from nitrogen.engine.contexts import ExcelContext
from nitrogen.engine.query import (
    DeleteQuery,
    FetchQuery,
    Filter,
    InsertQuery,
    Operator,
    UpdateQuery,
)
from openpyxl.worksheet.worksheet import Worksheet

class ExcelCompiler(QueryCompiler[ExcelContext]):
    def compile(self, query, context):
        if isinstance(query, FetchQuery):
            return self._compile_fetch(query, context.worksheet)
        elif isinstance(query, InsertQuery):
            return self._compile_insert(query, context.worksheet)
        elif isinstance(query, UpdateQuery):
            return self._compile_update(query, context.worksheet)
        elif isinstance(query, DeleteQuery):
            return self._compile_delete(query, context.worksheet)
        else:
            raise ValueError("unsupported query type")
    
    def _compile_fetch(self, query: FetchQuery, worksheet: Worksheet):
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            return []
        
        headers = rows[0]
        records = []

        for row in rows[1:]:
            record = dict(zip(headers, row))
            records.append(record)
        
        for filter_ in query.filters:
            predicate = self._compile_filter(filter_)
            records = [record for record in records if predicate(record)]
        
        if query.fields:
            records = [{k: record[k] for k in query.fields if k in record} for record in records]
        
        if query.order_by:
            order_by = query.order_by
            records.sort(key=lambda r: r.get(order_by))
        
        return records

    def _compile_insert(self, query: InsertQuery, worksheet: Worksheet):
        if worksheet.max_row == 0:
            headers = list(query.registry.keys())
            worksheet.append(headers)
        else:
            headers = [cell.value for cell in worksheet[1]]

        row = [query.registry.get(str(column)) for column in headers]
        worksheet.append(row)

    def _compile_update(self, query: UpdateQuery, worksheet: Worksheet):
        if worksheet.max_row < 1:
            return

        headers = [cell.value for cell in worksheet[1]]

        for row in worksheet.iter_rows(min_row=2):
            record = {headers[i]: row[i].value for i in range(len(headers))}
            record = cast(dict[str, Any], record)

            if all(self._compile_filter(f)(record) for f in query.filters):
                for i, column in enumerate(headers):
                    if column in query.registry:
                        row[i].value = query.registry[str(column)]
    
    def _compile_delete(self, query: DeleteQuery, worksheet: Worksheet):
        if worksheet.max_row < 1:
            return

        headers = [cell.value for cell in worksheet[1]]
        rows_to_delete = []

        for index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            record = {headers[i]: row[i].value for i in range(len(headers))}
            record = cast(dict[str, Any], record)

            if all(self._compile_filter(f)(record) for f in query.filters):
                rows_to_delete.append(index)
        
        for index in reversed(rows_to_delete):
            worksheet.delete_rows(index)
    
    def _compile_filter(self, filter: Filter) -> Callable[[dict[str, Any]], bool]:
        def resolve(record):
            return record.get(filter.field)

        match filter.operator:
            case Operator.EQ:
                return lambda r: resolve(r) == filter.value
            case Operator.GT:
                return lambda r: resolve(r) is not None and resolve(r) > filter.value
            case Operator.LT:
                return lambda r: resolve(r) is not None and resolve(r) < filter.value
            case Operator.GTE:
                return lambda r: resolve(r) is not None and resolve(r) >= filter.value
            case Operator.LTE:
                return lambda r: resolve(r) is not None and resolve(r) <= filter.value
            case Operator.CONTAINS:
                return lambda r: filter.value in (resolve(r) or "")
            case _:
                raise ValueError("unsupported operator")
    
    def _compile_group_by(self, records: list[dict], field: str):
        groups = defaultdict(list)

        for row in records:
            groups[row[field]].append(row)
        
        return dict(groups)
