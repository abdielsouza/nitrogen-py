import pytest
import sys
from pathlib import Path
from tempfile import NamedTemporaryFile
from openpyxl import Workbook, load_workbook

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from nitrogen.core import Column, Formula, Sheet
from nitrogen.core.expressions.base import Expression
from nitrogen.core.expressions.references import BasicReference
from nitrogen.backends.excel.backend import ExcelBackend


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def workbook():
    """Create a fresh Workbook for each test."""
    return Workbook()


@pytest.fixture
def backend(workbook):
    """Create an ExcelBackend with a fresh workbook."""
    return ExcelBackend(workbook)


@pytest.fixture
def simple_sheet():
    """Define a simple Sheet class with columns."""
    class SimpleSheet(Sheet):
        name = Column(dtype=str, name="name")
        age = Column(dtype=int, name="age")
    
    return SimpleSheet


@pytest.fixture
def sheet_with_data():
    """Define a Sheet with columns and pre-inserted data rows."""
    class DataSheet(Sheet):
        name = Column(dtype=str, name="name")
        age = Column(dtype=int, name="age")
    
    # Ensure clean state
    DataSheet.__rows__ = []
    
    # Insert sample data
    DataSheet.insert(name="Alice", age=30)
    DataSheet.insert(name="Bob", age=25)
    
    return DataSheet


# ============================================================================
# Tests: create_sheet + write_column (via integration test with saved file)
# ============================================================================

def test_create_sheet_and_write_columns(backend, simple_sheet):
    """Test creating a sheet and writing column headers."""
    backend.create_sheet(simple_sheet)
    backend.write_column(simple_sheet, simple_sheet.name)
    backend.write_column(simple_sheet, simple_sheet.age)
    
    # Save to temporary file and reload to verify
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        backend.save(tmp_path)
        
        # Load the saved file and verify structure
        loaded_wb = load_workbook(tmp_path)
        assert simple_sheet.__name__ in loaded_wb.sheetnames
        
        ws = loaded_wb[simple_sheet.__name__]
        assert ws["A1"].value == "name"
        assert ws["B1"].value == "age"
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_write_column_with_data(backend, sheet_with_data):
    """Test writing column headers and data rows."""
    backend.create_sheet(sheet_with_data)
    backend.write_column(sheet_with_data, sheet_with_data.name)
    backend.write_column(sheet_with_data, sheet_with_data.age)
    
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        backend.save(tmp_path)
        
        # Load and verify
        loaded_wb = load_workbook(tmp_path)
        ws = loaded_wb[sheet_with_data.__name__]
        
        # Check headers
        assert ws["A1"].value == "name"
        assert ws["B1"].value == "age"
        
        # Check first data row (row 2)
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 30
        
        # Check second data row (row 3)
        assert ws["A3"].value == "Bob"
        assert ws["B3"].value == 25
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_write_formula_column(backend):
    """Test writing a formula column into an Excel sheet."""
    class SalesSheet(Sheet):
        quantity = Column(dtype=int, name="quantity")
        price = Column(dtype=float, name="price")
        total = Formula(quantity * price)

    SalesSheet.__rows__ = []
    SalesSheet.insert(quantity=2, price=10.5)
    SalesSheet.insert(quantity=3, price=5.0)

    backend.create_sheet(SalesSheet)
    backend.write_column(SalesSheet, SalesSheet.quantity)
    backend.write_column(SalesSheet, SalesSheet.price)
    backend.write_formula(SalesSheet, SalesSheet.total)

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        backend.save(tmp_path)

        loaded_wb = load_workbook(tmp_path, data_only=False)
        ws = loaded_wb[SalesSheet.__name__]

        assert ws["C1"].value == "total"
        assert ws["C2"].value == "=A2*B2"
        assert ws["C3"].value == "=A3*B3"
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_multiple_sheets(backend, simple_sheet, sheet_with_data):
    """Test creating and writing to multiple sheets."""
    backend.create_sheet(simple_sheet)
    backend.write_column(simple_sheet, simple_sheet.name)
    
    backend.create_sheet(sheet_with_data)
    backend.write_column(sheet_with_data, sheet_with_data.name)
    backend.write_column(sheet_with_data, sheet_with_data.age)
    
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        backend.save(tmp_path)
        
        loaded_wb = load_workbook(tmp_path)
        assert simple_sheet.__name__ in loaded_wb.sheetnames
        assert sheet_with_data.__name__ in loaded_wb.sheetnames
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)


# ============================================================================
# Tests: save
# ============================================================================

def test_save_with_path(backend, simple_sheet):
    """Test saving the workbook to a file."""
    backend.create_sheet(simple_sheet)
    backend.write_column(simple_sheet, simple_sheet.name)
    
    # Create a temporary file
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        backend.save(tmp_path)
        
        # Verify file exists and has content
        assert Path(tmp_path).exists()
        assert Path(tmp_path).stat().st_size > 0
    finally:
        # Cleanup
        Path(tmp_path).unlink(missing_ok=True)


def test_save_without_path_raises_error(backend, simple_sheet):
    """Test that save without path raises ValueError."""
    backend.create_sheet(simple_sheet)
    backend.write_column(simple_sheet, simple_sheet.name)
    
    with pytest.raises(ValueError, match="missing path"):
        backend.save()


def test_save_with_none_path_raises_error(backend, simple_sheet):
    """Test that save with None path raises ValueError."""
    backend.create_sheet(simple_sheet)
    backend.write_column(simple_sheet, simple_sheet.name)
    
    with pytest.raises(ValueError, match="missing path"):
        backend.save(None)


# ============================================================================
# Integration Tests
# ============================================================================

def test_full_workflow_with_data(backend):
    """Test a full workflow: create sheet, write columns, and save."""
    
    # Define a sheet with columns
    class SalesSheet(Sheet):
        product = Column(dtype=str, name="product")
        quantity = Column(dtype=int, name="quantity")
        unit_price = Column(dtype=float, name="unit_price")
    
    # Reset __rows__ to ensure clean state
    SalesSheet.__rows__ = []
    
    # Insert sample data
    SalesSheet.insert(product="Widget", quantity=5, unit_price=10.5)
    SalesSheet.insert(product="Gadget", quantity=3, unit_price=25.0)
    
    # Create sheet and write data
    backend.create_sheet(SalesSheet)
    backend.write_column(SalesSheet, SalesSheet.product)
    backend.write_column(SalesSheet, SalesSheet.quantity)
    backend.write_column(SalesSheet, SalesSheet.unit_price)
    
    # Save to file
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        backend.save(tmp_path)
        
        # Verify file exists
        assert Path(tmp_path).exists()
        
        # Load and verify content
        loaded_wb = load_workbook(tmp_path)
        ws = loaded_wb[SalesSheet.__name__]
        
        # Verify headers
        assert ws["A1"].value == "product"
        assert ws["B1"].value == "quantity"
        assert ws["C1"].value == "unit_price"
        
        # Verify data rows
        assert ws["A2"].value == "Widget"
        assert ws["B2"].value == 5
        assert ws["C2"].value == 10.5
        
        assert ws["A3"].value == "Gadget"
        assert ws["B3"].value == 3
        assert ws["C3"].value == 25.0
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_empty_sheet_workflow(backend):
    """Test workflow with a sheet that has no data rows."""
    
    class EmptySheet(Sheet):
        name = Column(dtype=str, name="name")
        age = Column(dtype=int, name="age")
    
    # Ensure clean state
    EmptySheet.__rows__ = []
    
    backend.create_sheet(EmptySheet)
    backend.write_column(EmptySheet, EmptySheet.name)
    backend.write_column(EmptySheet, EmptySheet.age)
    
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        backend.save(tmp_path)
        
        loaded_wb = load_workbook(tmp_path)
        ws = loaded_wb[EmptySheet.__name__]
        
        # Headers should be present
        assert ws["A1"].value == "name"
        assert ws["B1"].value == "age"
        # But no data rows
        assert ws["A2"].value is None
        assert ws["B2"].value is None
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)
